import os
import tempfile
import traceback
import uuid

import flask
import flask_login
import flask_socketio
import openpyxl
import openpyxl.utils
from openpyxl.formatting import rule

from mathsnuggets import widgets
from mathsnuggets.core import cache, fields, form, models

api = flask.Blueprint("api", __name__)

socketio = flask_socketio.SocketIO(cors_allowed_origins="*")

widget_names = [n for n in dir(widgets) if n[0].isupper() and n[1].islower()]
widget_data = [{"path": n, "name": getattr(widgets, n).__doc__} for n in widget_names]


def is_post():
    return flask.request.method == "POST"


@api.route("/widgets")
@cache.cached()
def form_list():
    return flask.jsonify(widget_data)


@api.route("/surveys/<survey>", methods=["GET"])
def count_votes(survey):
    return flask.jsonify([dict(s) for s in models.Vote.find({"survey": survey})])


@api.route("/surveys/<survey>", methods=["DELETE"])
@flask_login.login_required
def delete_votes(survey):
    for vote in models.Vote.find({"survey": survey}):
        vote.delete()
    return flask.jsonify({"success": True})


@api.route("/surveys/<survey>", methods=["POST"])
def cast_vote(survey):
    user = str(flask.request.cookies.get("voter_id"))
    vote = models.Vote(user=user, survey=survey)
    if not vote.survey:
        vote.user = user
        vote.survey = survey
        vote.edit_count = -1
    vote.edit_count += 1
    payload = flask.request.get_json()
    if not vote.value and vote.edit_count <= payload["maxAttempts"]:
        vote.update(payload)
        socketio.emit(
            "voteReceived",
            {"user": str(vote.user), "value": vote.value, "survey": survey},
            room=survey,
        )
    return flask.jsonify(dict(vote))


@api.route("/mark", methods=["POST"])
def mark():
    payload = flask.request.get_json()
    form = models.Form(email=payload["email"], url=payload["url"])
    if form.email:
        raise InvalidUsage("Form already submitted for " + form.email)

    def parse_form(node, marks):
        if "children" in node:
            for child in node["children"]:
                parse_form(child, marks)
        elif node.get("component") == "widget":
            widget = getattr(widgets, node["type"])(**node["payload"])
            if len(dict(widget._fields(lambda f: f.get("nosave")))):
                marks.append(
                    {
                        "marks": getattr(widget, "_marks", 0),
                        "total": getattr(widget, "_total_marks", 1),
                    }
                )

    marks = []
    parse_form({"children": payload["form"]}, marks)
    form.update(
        {
            "children": payload["form"],
            "email": payload["email"],
            "first_name": payload["firstName"],
            "last_name": payload["lastName"],
            "marks": marks,
            "url": payload["url"],
            "year": payload["year"],
        }
    )
    score, total = 0, 0
    for m in marks:
        score += m["marks"]
        total += m["total"]
    return flask.jsonify([score, total])


@api.route("excel/<path:url>", methods=["GET"])
def generate_excel(url):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row, f in enumerate(models.Form.find({"url": url}), 1):
        fields = ["email", "first_name", "last_name", "year"]
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = getattr(f, field)
        for col, question in enumerate(f.marks, len(fields) + 1):
            bar = rule.DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=question["total"],
                color=openpyxl.styles.colors.BLUE,
            )
            letter = openpyxl.utils.get_column_letter(col)
            ws.conditional_formatting.add(f"{letter}{row}", bar)
            cell = ws.cell(row=row, column=col)
            cell.value = question["marks"]
    with tempfile.NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    return flask.Response(
        stream,
        headers={
            "Content-Disposition": "attachment; filename=sheet.xlsx",
            "Content-type": "application/vnd.openxmlformats-officedocument."
            + "spreadsheetml.sheet",
        },
    )


@socketio.on("join")
def join(survey):
    flask_socketio.join_room(survey)


@api.route("/slideshows", methods=["GET"])
@cache.memoize()
def list_slideshows():
    return flask.jsonify([dict(s) for s in models.Slideshow.find({})])


@api.route("/slideshows/<teacher>/<year>/<slug>", methods=["GET"])
@cache.memoize()
def load_slideshow(teacher=False, year=False, slug=False):
    params = {"teacher": teacher, "year": year, "slug": slug}
    return flask.jsonify(models.Slideshow(**params).children)


@api.route("/slideshows/", methods=["POST"])
@api.route("/slideshows/<teacher>/<year>/<slug>", methods=["POST"])
@flask_login.login_required
def save_slideshow(teacher=False, year=False, slug=False):
    params = {"teacher": teacher, "year": year, "slug": slug}
    slideshow = models.Slideshow(**params)
    slideshow.update(flask.request.get_json())
    cache.delete_memoized(list_slideshows)
    cache.delete_memoized(load_slideshow, teacher, year, slug)
    return flask.jsonify(dict(slideshow))


@api.route("/slideshows/<teacher>/<year>/<slug>", methods=["DELETE"])
@flask_login.login_required
def delete_slideshow(teacher=False, year=False, slug=False):
    params = {"teacher": teacher, "year": year, "slug": slug}
    slideshow = models.Slideshow(**params)
    slideshow.delete()
    cache.delete_memoized(list_slideshows)
    cache.delete_memoized(load_slideshow, teacher, year, slug)
    return "", 204


@api.route("/fields/<field>", methods=["GET"])
@cache.cached(query_string=True)
def validate_field(field):
    payload = flask.request.args
    if not hasattr(fields, field):
        raise InvalidUsage(f"Field {repr(field)} does not exist", 404, payload=payload)
    if "value" not in payload:
        raise InvalidUsage("Missing 'value' field in payload", payload=payload)
    field_cls = getattr(fields, field)
    value = payload.get("value")
    kwargs = {k: v for k, v in payload.items() if k not in field_cls._blacklist}

    class DummyForm(form.Form):
        field = field_cls("Dummy Field", **kwargs)

    try:
        response = DummyForm.field.validate(value)
        return flask.jsonify(response)
    except (ValueError, TypeError, AttributeError) as error:
        raise InvalidUsage(str(error), 400, payload)


@api.route("/widgets/<path:widget>", methods=["GET", "POST"])
@cache.cached(unless=is_post, query_string=True)
def form_route(widget):
    payload = flask.request.args or flask.request.get_json()
    if widget not in widget_names:
        raise InvalidUsage(f"Widget {repr(widget)} does not exist", 404, payload)
    try:
        widget = getattr(widgets, widget)(**(payload if payload else {}))
        if is_post():
            widget.generate()
        if payload:
            widget._validate()
            return flask.jsonify(dict(widget))
    except (AttributeError, ValueError, TypeError, PermissionError) as error:
        raise InvalidUsage(str(error), 400, payload)
    return flask.jsonify(
        {
            "template": str(widget),
            "fields": dict(widget._fields()),
            "generator_template": widget._template(
                getattr(widget, "generator_template", "")
            ),
            "constraints": widget.get_fields_as_template("constraint"),
            "random_numbers": widget.get_fields_as_template("random"),
        }
    )


login_manager = flask_login.LoginManager()


@login_manager.user_loader
def load_user(identifier):
    user = models.User(_id=identifier)
    return user if user.email else None


@api.route("/auth/register", methods=["POST"])
def register():
    payload = flask.request.get_json()
    registration_pw = payload.pop("registration_password")
    if registration_pw != os.environ.get("REGISTRATION_PASSWORD", ""):
        raise InvalidUsage("Registration password incorrect", 400, payload)
    user = models.User(email=payload["email"])
    if user.email:
        raise InvalidUsage(f"User {user.email} already exists", 400, payload)
    try:
        user.update(payload)
    except (AttributeError, ValueError) as error:
        raise InvalidUsage(str(error), 400, payload)
    return flask.jsonify({"success": True})


@api.route("/auth/login", methods=["GET", "POST"])
def login():
    payload = flask.request.get_json()
    if not payload:
        user = flask_login.current_user
        voter_id = flask.request.cookies.get("voter_id") or str(uuid.uuid1())
        response = flask.jsonify(
            {
                "is_authenticated": user.is_authenticated,
                "email": getattr(user, "email", False),
                "voter_id": voter_id,
            }
        )
        if not flask.request.cookies.get("voter_id"):
            response.set_cookie("voter_id", voter_id)
        return response
    user = models.User(email=payload["email"])
    if not user.email or not user.check_password(payload["password"]):
        raise InvalidUsage("Incorrect email or password")
    flask_login.login_user(user, remember=payload.get("remember", False))
    return flask.jsonify({"success": True})


@api.route("/auth/nickname", methods=["GET", "POST"])
def update_nickname():
    payload = flask.request.get_json()
    anonymous_id = flask.request.cookies.get("voter_id")
    username = getattr(flask_login.current_user, "email", anonymous_id)
    identity = models.Identity(username=username)
    if payload:
        payload["username"] = username
        identity.update(payload)
    return flask.jsonify(dict(identity))


@api.route("/auth/logout")
@flask_login.login_required
def logout():
    flask_login.logout_user()
    return flask.jsonify({"success": True})


@api.route("/tests")
@cache.cached()
def tests():
    test_data = {v["name"]: v["test"] for k, v in widgets.info.items() if v.get("test")}
    return flask.jsonify(test_data)


@api.route("/tests/user", methods=["DELETE"])
def delete_test_user():
    if "MONGO_URL" not in os.environ:
        user = models.User(email="test@test.com")
        user.delete()
        return "", 204


class InvalidUsage(Exception):
    status_code = 400

    def __init__(self, message, status_code=None, payload=None):
        super().__init__(self)
        self.message = message
        if status_code is not None:
            self.status_code = status_code
        self.payload = payload

    def __iter__(self):
        yield ("error", True)
        yield ("status_code", self.status_code)
        yield ("traceback", traceback.format_exc())
        yield ("payload", dict(self.payload or ()))
        yield ("message", self.message)


@api.errorhandler(InvalidUsage)
def handle_invalid_usage(error):
    response = flask.jsonify(dict(error))
    response.status_code = error.status_code
    return response
